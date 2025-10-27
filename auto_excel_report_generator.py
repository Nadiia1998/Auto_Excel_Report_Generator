import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference , LineChart
from openpyxl import Workbook
import calendar
import os
import sys



# Get path of the running script

application_path = os.path.dirname(sys.executable)

data_file = os.path.join(application_path, 'ecommerce.csv')

df= pd.read_csv(data_file)

df['Order_Date'] = pd.to_datetime(df['Order_Date'], errors='coerce')
df['Delivery_Date'] = pd.to_datetime(df['Delivery_Date'], errors='coerce')
df['Delivery_Duration'] = (df['Delivery_Date'] - df['Order_Date']).dt.days

def delivery_speed(days):
    if days <=3:
        return 'Fast'
    elif days <= 7:
        return "Medium"
    else: return "Slow"
df['Delivery_Category']= df['Delivery_Duration'].apply(delivery_speed)

df['Price'] = df['Price'].str.replace(',','.').astype('float')
df['Total_Value'] = df['Total_Value'].str.replace(',','.').astype('float')


df_2023 = df[df['Order_Date'].dt.year == 2023]




# Create a new workbook and set the active sheet
wb=Workbook()
ws = wb.active
ws.title = "Monthly Report"
min_column = ws.min_column


# --- User input for month ---
month_input = int(input("Enter month number (1-12): "))
df_month = df_2023[df_2023['Order_Date'].dt.month == month_input]
month_name= calendar.month_name[month_input]


# --- Calculate main KPIs ---
total_order = df_month['Order_ID'].nunique()
total_revenu = df_month['Total_Value'].sum().round(2)
AOV = df_month['Total_Value'].mean().round(2)
total_quality_sold = df_month['Quantity'].sum().round(0)

if df_month.empty:
    print("No data for this month!")
    exit()

# --- Styling for header and KPI cells ---
major_color = PatternFill(start_color="FF4D587A", end_color="FF4D587A", fill_type="solid")
minor_color =PatternFill(start_color="FF6D8FAA", end_color="FF6D8FAA", fill_type="solid")

# Header
ws['A1'].fill = major_color
ws.merge_cells('A1:T1')
header = ws['A1']
header.value = f'E-Commerce Report - Month {month_name}'
header.font = Font(size=10, bold = True, color ='FFFFFFFF')
header.alignment = Alignment(horizontal = 'center')


# Merge cells for KPI display
merge_cells = ['B3:D3', 'G3:I3','L3:N3','Q3:S3','B4:D4', 'G4:I4','L4:N4','Q4:S4']
cell_add = ['B','G','L','Q']

for i in merge_cells:
    ws.merge_cells(i)


# KPI labels and values
kpis=[
    ('Total Revenue',total_revenu),
    ('Average Order Value',AOV),
    ('Total Quantity',total_quality_sold),
    ('Total Order', total_order)
]


for i, (name, value) in enumerate(kpis):
    cell = cell_add[i]
    ws[f'{cell}{3}'] = name
    ws[f'{cell}{4}'] = value
    ws[f'{cell}{3}'].font = Font(bold=True, color = 'FFFFFFFF' )
    ws[f'{cell}{4}'].number_format = '#,##0.00' if isinstance(value,float) else '#,##0'
    ws[f'{cell}{3}'].fill = minor_color
    ws[f'{cell}{4}'].alignment = Alignment(horizontal = 'center')


# --- Sales by Category Table ---
agg_category = df_month.groupby('Product_Category').agg(total_value = ('Total_Value','sum'),
                                       average_value = ('Total_Value','mean'),
                                        quantity = ('Quantity','sum')).round(2).reset_index().sort_values('total_value',ascending = False)

columns = list(agg_category.columns)

# Add table headers
for col_num, column_title in enumerate(columns, 7):
    cell = ws.cell(row=22, column=col_num)
    cell.value = column_title
    cell.font = Font(size=14 ,bold=True, color ='FFFFFFFF')
    cell.fill = minor_color
    cell.alignment = Alignment(horizontal='center')

# Add table data
for row_num, row_data in enumerate(agg_category.values, 23):
    for col_num, cell_value in enumerate(row_data, 7):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = Font(size = 12)
        if isinstance(cell_value, (int, float)) and cell_value > 10000:
            cell.number_format = '$#,##0'

# Add borders to the table
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

for row in ws.iter_rows(min_row=22, max_row=30, min_col=7, max_col=10):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


# --- Sales by Country Bar Chart ---
sale_per_country = df_month.groupby('Shipping_Country')['Total_Value'].sum().reset_index().sort_values(by='Total_Value',ascending=False)
categories = list(sale_per_country['Shipping_Country'])
sales = list(sale_per_country['Total_Value'])

for i, (cat, val) in enumerate(zip(categories, sales), start=6):
    ws[f'B{i}'] = cat
    ws[f'C{i}'] = val


chart = BarChart()
chart.styles = 5
chart.title = "Sales by Country"
chart.x_axis.title = "Country"
chart.y_axis.title = "Total Sales"

chart.y_axis.majorGridlines = None
chart.x_axis.majorGridlines = None


chart.x_axis.tickLblPos = "low"
chart.y_axis.tickLblPos = "nextTo"

data = Reference(ws, min_col=3,min_row=6, max_row=13)
categories_ref = Reference(ws, min_col=2, min_row=6, max_row=13)
chart.add_data(data, titles_from_data=False)
chart.set_categories(categories_ref)

ws.add_chart(chart, "B6")



# --- Daily Sales Line Chart ---
day_sale = df_month.groupby('Order_Date',as_index=False)['Total_Value'].sum()

category_line = list (day_sale['Order_Date'])
value_line = list (day_sale['Total_Value'])

for i, (cat,val) in enumerate(zip(category_line,value_line),start=6):
    ws[f'X{i}'] = cat
    ws[f'Y{i}'] = val

linechar = LineChart()
linechar.styles = 5
linechar.title = "Sales Trend"
#linechar.x_axis = DateAxis()
linechar.x_axis.title = "Day"
linechar.y_axis.title = "Total Sales"

linechar.y_axis.majorGridlines = None
linechar.x_axis.majorGridlines = None


linechar.y_axis.tickLblPos = "nextTo"


line_data =Reference(ws,min_col= 25, min_row =6, max_row= 37)
line_category = Reference(ws, min_col= 24, min_row=6, max_row=37)

linechar.add_data(line_data,titles_from_data=True)
linechar.set_categories(line_category)

linechar.legend = None
s1 = linechar.series[0]
s1.graphicalProperties.line.solidFill = "4472C4"  # синій
s1.graphicalProperties.line.width = 20000  # товщина лінії
ws.add_chart(linechar, "L6")


# --- Delivery Information ---
delivery = df_month['Delivery_Category'].value_counts().reset_index()
delivery.columns = ['Category','Value']

delivery_categories = list(delivery['Category'])
delivery_value = list(delivery['Value'])

for i, (cat, val) in enumerate(zip(delivery_categories, delivery_value), start=22):
    ws[f'B{i}'] = cat
    ws[f'C{i}'] = val

# Delivery Bar Chart
deliver_char = BarChart()

deliver_char.styles = 5
deliver_char.title = "Delivery Info"


deliver_char.y_axis.majorGridlines = None
deliver_char.x_axis.majorGridlines = None

deliver_char.x_axis.tickLblPos = "low"
deliver_char.y_axis.tickLblPos = "nextTo"


data = Reference(ws, min_col=3,min_row=22, max_row=24)
categories_ref = Reference(ws, min_col=2, min_row=22, max_row=24)
deliver_char.add_data(data, titles_from_data=False)
deliver_char.set_categories(categories_ref)

deliver_char.width = 7
deliver_char.height = 5
ws.add_chart(deliver_char, "B22")

# --- Pivot Table for Product Category vs Shipping Country ---
pivot_table = df_month.pivot_table(index='Product_Category', columns='Shipping_Country',values='Total_Value',aggfunc='sum').reset_index()
pivot_table


pivot_table_columns = list(pivot_table.columns)

for col_num, column_title in enumerate(pivot_table_columns, 12):
    cell = ws.cell(row=22, column=col_num)
    cell.value = column_title
    cell.font = Font(size=14 ,bold=True, color = 'FFFFFFFF')
    cell.fill =minor_color
    cell.alignment = Alignment(horizontal='center')


for row_num, row_data in enumerate(pivot_table.values, 23):
    for col_num, cell_value in enumerate(row_data, 12):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = Font(size = 12)
        cell.number_format = '$#,##0'

# Add pivot table data
for row in ws.iter_rows(min_row=22, max_row=30, min_col=12, max_col=20):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# --- Save Workbook ---
output_path = os.path.join(application_path, f'Report_Month_{month_name}.xlsx')
wb.save(output_path)
